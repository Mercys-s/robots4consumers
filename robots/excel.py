import openpyxl as excel


def info_to_excel(excel_file_path, v_data, db_model):
    robot_queryset = db_model.objects.all().values('model')
    robot_models = [robot["model"] for robot in robot_queryset]

    book = excel.open(excel_file_path)
    sheet = book.active
    worksheet_names = [worksheet.title for worksheet in book.worksheets]
    
    for model in robot_models:
        if str(model) not in worksheet_names:
            
            book.create_sheet(title = model)
            book.save(excel_file_path)

    worksheet_objects = [worksheet for worksheet in book.worksheets]

    for worksheet in worksheet_objects:
        
        if worksheet.title == v_data["model"]:
            sheet_index = book.sheetnames.index(worksheet.title)
            sheet = book.worksheets[sheet_index]

            if sheet['A1'].value or sheet['B1'].value or sheet['C1'].value == None:
                sheet['A1'].value = "Модель"
                sheet['B1'].value = "Серия"
                sheet['C1'].value = "Кол-во за неделю"

            if sheet.title == v_data["model"]: 
                adding_model_info = [v_data["model"], v_data["version"]]
                is_match = match_check(adding_model_info, sheet)

                if is_match == None:
                    sheet.append([v_data["model"], v_data["version"], 1])

                if is_match != None:
                    sheet[is_match][2].value = int(sheet[is_match][2].value) + 1

            book.save(excel_file_path)
            book.close()



def match_check(adding_model, sheet):

    for row in range(1, sheet.max_row + 1):

        model = sheet[row][0].value
        serial = sheet[row][1].value
        row_info = [model, serial]

        if adding_model == row_info:
            return row